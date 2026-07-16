"""Test the universal document processing system with synthetic files in unusual layouts."""
import io
import sys
import datetime as dt

sys.path.insert(0, ".")

import pandas as pd
from openpyxl import Workbook
from modules import universal_parser as up
from modules import data_processor as dp

PASS, FAIL = 0, 0


def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {name}")
    else:
        FAIL += 1
        print(f"  FAIL  {name}  {detail}")


# ---------------------------------------------------------------- 1. CSV, odd headers, messy numbers
print("\n[1] CSV with unfamiliar headers + ₹/commas/parens")
csv_text = """Txn Dt,Airport,Business Line,Unit,Guests,Sales (INR)
05/07/2026,DEL,Lounge,Encalm Lounge T3 Intl,1250,"₹ 12,50,000.50"
05/07/2026,HYD,Lounge,Encalm Lounge RGIA Dom,890,"Rs. 8,90,450"
05/07/2026,GOI,Meet & Greet,Atithya Goa,45,"₹ 2,25,000"
05/07/2026,DEL,F&B,Encalm Eats T3,300,"(15,000)"
Grand Total,,,,2485,"₹ 23,50,450"
"""
r = up.parse_universal(io.BytesIO(csv_text.encode()), "daily_sales.csv")
df = r.df
check("parsed rows (total row excluded)", len(df) == 4, f"got {len(df)}")
check("date parsed day-first", df["date"].iloc[0] == dt.date(2026, 7, 5), df["date"].iloc[0])
check("DEL -> Delhi", set(df["location"]) == {"Delhi", "Hyderabad", "Goa"}, set(df["location"]))
check("₹+lakh-style commas parsed", abs(df["revenue"].iloc[0] - 1250000.50) < 0.01, df["revenue"].iloc[0])
check("parens = negative", df["revenue"].iloc[3] == -15000, df["revenue"].iloc[3])
check("guests -> pax", df["pax"].iloc[0] == 1250, df["pax"].iloc[0])
check("segment matched", df["segment"].iloc[2] == "Atithya", df["segment"].iloc[2])
check("confidence >= threshold", r.confidence >= 0.55, r.confidence)
print("   mapping:", "; ".join(f"{m.role}<-{m.source}" for m in r.mappings))

# ---------------------------------------------------------------- 2. Excel: title rows, date in title, section-label locations, no segment column
print("\n[2] Excel: no Date column (date in title), Location as section labels, segment derived from outlet names")
wb = Workbook(); ws = wb.active; ws.title = "Report"
ws.append(["Encalm Group — Daily Performance Report"])
ws.append(["For 06-Jul-2026"])
ws.append([])
ws.append(["Station", "Outlet Name", "Footfall", "Amount"])
ws.append(["Delhi", "Encalm Lounge T3", 1500, 1600000])
ws.append([None, "Encalm Spa T3", 120, 240000])
ws.append([None, "Sky Plates Kitchen", 0, 95000])
ws.append(["Hyderabad", "Encalm Lounge RGIA", 700, 720000])
ws.append([None, "Atithya Services HYD", 60, 180000])
ws.append(["Total", "", 2380, 2835000])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
r = up.parse_universal(buf, "performance.xlsx")
df = r.df
check("rows", len(df) == 5, f"got {len(df)}\n{df}")
check("date recovered from title", (df["date"] == dt.date(2026, 7, 6)).all(), df["date"].unique())
check("location forward-filled", list(df["location"]) == ["Delhi", "Delhi", "Delhi", "Hyderabad", "Hyderabad"], list(df["location"]))
segs = list(df["segment"])
check("segments derived from outlet keywords",
      segs == ["Lounges", "Others", "Subsidiary", "Lounges", "Atithya"], segs)

# ---------------------------------------------------------------- 3. Excel wide-by-date (dates as columns)
print("\n[3] Excel: wide layout, dates as columns")
wb = Workbook(); ws = wb.active; ws.title = "Delhi"
ws.append(["Revenue Report — Delhi Airport"])
ws.append(["Segment", "Outlet", "01-07-2026", "02-07-2026", "03-07-2026"])
ws.append(["Lounges", "Encalm Lounge T3", 1000000, 1100000, 1050000])
ws.append(["Atithya", "Atithya T3", 200000, 210000, 190000])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
r = up.parse_universal(buf, "delhi_wide.xlsx")
df = r.df
check("melted to 6 rows", len(df) == 6, f"got {len(df)}\n{df}")
check("location from sheet name", (df["location"] == "Delhi").all(), df["location"].unique())
check("3 distinct dates", df["date"].nunique() == 3, df["date"].unique())
check("melted values are revenue", df["revenue"].sum() == 3750000, df["revenue"].sum())

# ---------------------------------------------------------------- 4. PDF with a generic (non-Encalm) table
print("\n[4] PDF: generic table not matching the predefined Encalm layout")
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table
pdf_buf = io.BytesIO()
doc = SimpleDocTemplate(pdf_buf, pagesize=A4)
rows = [["Date", "City", "Business", "Outlet", "Passengers", "Revenue"],
        ["05-07-2026", "Delhi", "Lounges", "Encalm Lounge T1", "400", "4,00,000"],
        ["05-07-2026", "Goa", "Others", "Encalm Spa Goa", "50", "1,10,000"]]
doc.build([Table(rows)])
pdf_buf.seek(0)
r = up.parse_universal(pdf_buf, "generic_report.pdf")
df = r.df
check("pdf rows", len(df) == 2, f"got {len(df)}\n{df}")
check("pdf revenue parsed with Indian commas", df["revenue"].iloc[0] == 400000, df["revenue"].iloc[0])
check("pdf locations", set(df["location"]) == {"Delhi", "Goa"}, set(df["location"]))

# ---------------------------------------------------------------- 5. AOP guard
print("\n[5] Guard: AOP/budget workbook must be REJECTED as revenue actuals")
wb = Workbook(); ws = wb.active; ws.title = "AOP FY27"
ws.append(["Annual Operating Plan — Budget FY 2026-27 (Rs. in Lakhs)"])
ws.append(["Outlet", "Location", "01-04-2026", "01-05-2026", "01-06-2026"])
ws.append(["Encalm Lounge T3", "Delhi", 250, 260, 270])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
try:
    up.parse_universal(buf, "AOP_Budget_FY27.xlsx")
    check("AOP file rejected", False, "was accepted!")
except up.UniversalParseError as e:
    check("AOP file rejected", "AOP" in str(e) or "budget" in str(e).lower(), str(e)[:100])

# ---------------------------------------------------------------- 6. Traffic-only file must fail universal (no revenue)
print("\n[6] Guard: traffic-only file fails universal (so the traffic pipeline can claim it)")
wb = Workbook(); ws = wb.active
ws.append(["Date", "Location", "Traffic"])
ws.append([dt.date(2026, 7, 5), "Delhi", 195000])
ws.append([dt.date(2026, 7, 5), "Goa", 28000])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
try:
    up.parse_universal(buf, "airport_traffic.xlsx")
    check("traffic-only rejected by universal", False, "was accepted!")
except up.UniversalParseError:
    check("traffic-only rejected by universal", True)

# ---------------------------------------------------------------- 7. Full pipeline via data_processor (no DB write)
print("\n[7] data_processor integration (save_to_db=False)")
res = dp.process_uploaded_file(io.BytesIO(csv_text.encode()), "daily_sales.csv", save_to_db=False)
check("csv through process_uploaded_file", res.success, res.message)
check("schema report surfaced in warnings", any("Auto-detected schema" in w for w in res.warnings),
      res.warnings[:2])
check("canonical columns present", res.df is not None and
      set(["date", "segment", "outlet", "location", "pax", "revenue"]).issubset(res.df.columns))

# unusual Excel through the SAME entry point Home.py uses (fallback path)
buf2 = io.BytesIO()
wb = Workbook(); ws = wb.active; ws.title = "Report"
ws.append(["Encalm Group — Daily Performance Report"])
ws.append(["For 06-Jul-2026"])
ws.append([])
ws.append(["Station", "Outlet Name", "Footfall", "Amount"])
ws.append(["Delhi", "Encalm Lounge T3", 1500, 1600000])
ws.append(["Hyderabad", "Encalm Lounge RGIA", 700, 720000])
wb.save(buf2); buf2.seek(0)
res = dp.process_uploaded_file(buf2, "performance.xlsx", save_to_db=False)
check("unknown-layout Excel falls back to universal", res.success, res.message)
check("fallback is flagged to the user",
      any("did not match any predefined report layout" in w for w in res.warnings), res.warnings[:1])

# ---------------------------------------------------------------- 8. Regression: predefined long-format Excel still uses the predefined parser
print("\n[8] Regression: predefined long-format workbook still parses (predefined path, not universal)")
wb = Workbook(); ws = wb.active; ws.title = "Data"
ws.append(["Date", "Location", "Business", "Sub-Business", "PAX", "Revenue"])
for d in [dt.date(2026, 7, 1), dt.date(2026, 7, 2)]:
    ws.append([d, "Delhi", "Lounges", "Encalm Lounge T3", 1000, 1000000])
    ws.append([d, "Goa", "Others", "Encalm Spa Goa", 40, 80000])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
res = dp.process_uploaded_file(buf, "Revenue_Dashboard.xlsx", save_to_db=False)
check("predefined format still succeeds", res.success, res.message)
check("did NOT go through universal fallback",
      not any("Auto-detected schema" in w for w in res.warnings), res.warnings)

# ---------------------------------------------------------------- 9. Auto-router: traffic file still routes to traffic pipeline
print("\n[9] Auto-router still sends a traffic file to the traffic pipeline")
wb = Workbook(); ws = wb.active
ws.append(["Date", "Location", "Traffic"])
ws.append([dt.date(2026, 7, 5), "Delhi", 195000])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
res = dp.process_uploaded_file_auto(buf, "traffic.xlsx", save_to_db=False)
check("routed as traffic", res.success and "traffic" in res.message.lower(), res.message)

# ---------------------------------------------------------------- 10. Units multiplier from title
print("\n[10] 'Rs. in Lakhs' title multiplies revenue")
wb = Workbook(); ws = wb.active; ws.title = "Delhi"
ws.append(["Daily Revenue (Rs. in Lakhs) — 05-07-2026"])
ws.append(["Segment", "Outlet", "Guests", "Revenue"])
ws.append(["Lounges", "Encalm Lounge T3", 1200, 12.5])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)
r = up.parse_universal(buf, "lakhs.xlsx")
check("lakhs multiplier applied", r.df["revenue"].iloc[0] == 1250000, r.df["revenue"].iloc[0])
check("pax NOT multiplied", r.df["pax"].iloc[0] == 1200, r.df["pax"].iloc[0])

print(f"\n===== {PASS} passed, {FAIL} failed =====")
sys.exit(1 if FAIL else 0)
